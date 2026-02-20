import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.db.models import AuctionListing, UsedCarListing


def _create_header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
        "align": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def _add_stats_sheet(wb, stats):
    ws_stats = wb.create_sheet("Statistics")
    ws_stats.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws_stats.cell(row=1, column=2, value="Value").font = Font(bold=True)
    for i, (key, value) in enumerate(stats.items(), 2):
        label = key.replace("_", " ").title()
        ws_stats.cell(row=i, column=1, value=label)
        ws_stats.cell(row=i, column=2, value=value)
    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 15


def export_listings_to_excel(listings: list[AuctionListing], stats: dict | None = None) -> io.BytesIO:
    """Generate an Excel file from auction listings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auction Results"

    style = _create_header_style()
    headers = [
        "Year", "Make", "Model", "Starting Bid", "Sold Price",
        "Auction Days", "Bids", "Times Listed", "Status", "Platform", "Link"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["align"]
        cell.border = style["border"]

    money_fmt = '#,##0'
    light_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

    for row_idx, listing in enumerate(listings, 2):
        ws.cell(row=row_idx, column=1, value=listing.year)
        ws.cell(row=row_idx, column=2, value=listing.make)
        ws.cell(row=row_idx, column=3, value=listing.model)

        starting_cell = ws.cell(row=row_idx, column=4, value=listing.starting_bid)
        starting_cell.number_format = money_fmt

        sold_cell = ws.cell(row=row_idx, column=5, value=listing.sold_price)
        sold_cell.number_format = money_fmt

        ws.cell(row=row_idx, column=6, value=listing.auction_days)
        ws.cell(row=row_idx, column=7, value=listing.bid_count)
        ws.cell(row=row_idx, column=8, value=listing.times_listed)
        ws.cell(row=row_idx, column=9, value="Sold" if listing.is_sold else "Not Sold")

        # Platform name (resolve from relationship or store)
        ws.cell(row=row_idx, column=10, value=getattr(listing, '_platform_name', ''))

        # Hyperlink to listing
        if listing.url:
            link_cell = ws.cell(row=row_idx, column=11, value="View Listing")
            link_cell.hyperlink = listing.url
            link_cell.font = Font(color="0563C1", underline="single")

        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = light_fill

    col_widths = [8, 12, 25, 14, 14, 13, 8, 13, 10, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if stats:
        _add_stats_sheet(wb, stats)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_used_cars_to_excel(listings: list[UsedCarListing], stats: dict | None = None) -> io.BytesIO:
    """Generate an Excel file from used car listings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Used Car Listings"

    style = _create_header_style()
    headers = [
        "Year", "Make", "Model", "Trim", "List Price", "Currency",
        "Mileage", "Days on Market", "Dealer", "Location", "Platform", "Link"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["align"]
        cell.border = style["border"]

    money_fmt = '#,##0'
    number_fmt = '#,##0'
    light_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

    for row_idx, listing in enumerate(listings, 2):
        ws.cell(row=row_idx, column=1, value=listing.year)
        ws.cell(row=row_idx, column=2, value=listing.make)
        ws.cell(row=row_idx, column=3, value=listing.model)
        ws.cell(row=row_idx, column=4, value=listing.trim)

        price_cell = ws.cell(row=row_idx, column=5, value=listing.list_price)
        price_cell.number_format = money_fmt

        ws.cell(row=row_idx, column=6, value=listing.currency or "USD")

        mileage_cell = ws.cell(row=row_idx, column=7, value=listing.mileage)
        mileage_cell.number_format = number_fmt

        ws.cell(row=row_idx, column=8, value=listing.days_on_market)
        ws.cell(row=row_idx, column=9, value=listing.dealer_name)
        ws.cell(row=row_idx, column=10, value=listing.location)

        ws.cell(row=row_idx, column=11, value=getattr(listing, '_platform_name', ''))

        if listing.url:
            link_cell = ws.cell(row=row_idx, column=12, value="View Listing")
            link_cell.hyperlink = listing.url
            link_cell.font = Font(color="0563C1", underline="single")

        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = light_fill

    col_widths = [8, 12, 20, 18, 14, 10, 12, 15, 20, 15, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if stats:
        _add_stats_sheet(wb, stats)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
